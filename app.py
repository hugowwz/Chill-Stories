from flask import Flask, request, render_template, stream_template, url_for, redirect
from flask_login import current_user, login_user, login_required, logout_user
from models import login, db, UserModel
import socket
# import logging
from ast import Break, While
from cProfile import label
from cmd import Cmd
from distutils import command
from email.mime import image
from multiprocessing import Condition
from tkinter import *
from tkinter.messagebox import showerror, showinfo
from tracemalloc import stop
from turtle import left, width
from wave import Wave_write
import tkinter.font as font
import os
import openpyxl

app = Flask(__name__)
app.secret_key = 'xyz'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///models.py'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
 
@app.route("/")
def welcome():
    return render_template('home.html')

db.init_app(app)
login.init_app(app)
login.login_view = 'login'

@app.before_first_request
def create_all():
    db.create_all()
     
@app.route('/home')
@login_required
def blog():
    return render_template('main.html')

@app.route('/login', methods = ['POST', 'GET'])
def login():
    if current_user.is_authenticated:
        return redirect('/home')
     
    if request.method == 'POST':
        email = request.form['email']
        user = UserModel.query.filter_by(email = email).first()
        password_hash = request.form['password']

        if email == "" and password_hash == "":
            error = "Complete all fields"
            return render_template("login.html", error=error)

        if email == "":
            error = "Please enter email"
            return render_template("login.html", error=error)

        if password_hash == "":
            error = "Please enter password"
            return render_template("login.html", error=error)

        if not UserModel.query.filter_by(email=email).first():
            error = "Email is not valid"
            return render_template("login.html", error=error)

        if user is not None and user.check_password(request.form['password']):
            login_user(user)
            app.logger.info("\x1b[31;20m" + "+------- USER HAS LOGGED IN -------+" + "\x1b[0m")
            app.logger.info("\x1b[31;20m" + "| Email    : " + "\x1b[0m" + email)
            app.logger.info("\x1b[31;20m" + "| Username : " + "\x1b[0m" + '%s', user.username)
            app.logger.info("\x1b[31;20m" + "| Password : " + "\x1b[0m" +  password_hash)
            app.logger.info("\x1b[31;20m" + "+" + "\x1b[0m")
            return redirect('/home')

        if not UserModel.query.filter_by(password_hash=password_hash).first():
            error = "Password is not valid"
            return render_template("login.html", error=error)
     
    return render_template('login.html')

@app.route('/signup', methods=['POST', 'GET'])
def register():
    if current_user.is_authenticated:
        return redirect('/home')
     
    if request.method == 'POST':
        email = request.form['email']
        username = request.form['username']
        password = request.form['password']
 
        if UserModel.query.filter_by(email=email).first():
            error = "Email already present"
            return render_template("signup.html", error=error)

        if email == "" and password == "" and username =="":
            error = "Complete all fields"
            return render_template("signup.html", error=error)

        if email == "":
            error = "Please enter email"
            return render_template("signup.html", error=error)

        if password == "":
            error = "Please enter password"
            return render_template("signup.html", error=error)

        if username =="":
            error = "Please enter username"
            return render_template("signup.html", error=error)
             
        user = UserModel(email=email, username=username)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        if user is not None and user.check_password(request.form['password']):
            login_user(user)
            app.logger.info("\x1b[31;20m" + "+---- NEW USER HAS REGISTERED ----+" + "\x1b[0m")
            app.logger.info("\x1b[31;20m" + "| Email    : " + "\x1b[0m" + email)
            app.logger.info("\x1b[31;20m" + "| Username : " + "\x1b[0m" + '%s', user.username)
            app.logger.info("\x1b[31;20m" + "| Password : " + "\x1b[0m" + password)
            app.logger.info("\x1b[31;20m" + "+" + "\x1b[0m")

            ############################################

            wb = openpyxl.load_workbook("C:\\Users\\Hugow\\Desktop\\Serveur\\static\\database\\bdd.xlsx")
            sheet = wb.active

            cpt = 5
            ctest = sheet.cell(row = cpt, column = 2)
            while ctest.value != None:
                cpt = cpt + 1
                ctest = sheet.cell(row = cpt, column = 2)

            c1 = sheet.cell(row = cpt, column = 2) 
            c1.value = cpt - 4
            c2 = sheet.cell(row= cpt, column = 3) 
            c2.value = email
            c3 = sheet.cell(row = cpt, column = 4)
            c3.value = user.username
            c4 = sheet.cell(row = cpt, column = 5)
            c4.value = password
            wb.save("C:\\Users\\Hugow\\Desktop\\Serveur\\static\\database\\bdd.xlsx")

            ############################################

        return redirect('/login')
    return render_template('signup.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect('/home')

@app.errorhandler(404)
def page_not_found(error):
    return render_template("error404.html"), 404

@app.route('/film/adam-a-travers-le-temps')
@login_required
def adam_a_travers_le_temps():
    return render_template('adam-a-travers-le-temps.html')

@app.route('/film/forces-speciales')
@login_required
def forces_speciales():
    return render_template('forces-speciales.html')

@app.route('/film/aquaman')
@login_required
def aquaman():
    return render_template('aquaman.html')

@app.route('/film/doctor-strange')
@login_required
def doctor_strange():
    return render_template('doctor-strange.html')

@app.route("/account")
@login_required
def account():
    return render_template('user-page.html')

ip = socket.gethostbyname(socket.gethostname())
if __name__ == "__main__":
    app.run(host=ip ,port=80,debug=True)